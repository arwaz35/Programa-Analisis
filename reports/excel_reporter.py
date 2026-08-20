"""
Generador de reportes Excel.
Usa plantillas .xlsx y las llena con datos e imágenes.
"""
import os
import io
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import pandas as pd
from PIL import Image
from config import FORMATOS_DIR, RESULTADOS_DIR, PILOTOS_FOTOS_DIR, MOTOS_FOTOS_DIR


class ExcelReporter:
    def __init__(self, templates_dir=None, output_dir=None):
        self.templates_dir = templates_dir or FORMATOS_DIR
        self.output_dir = output_dir or RESULTADOS_DIR

        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def _insert_image(self, ws, img_bytes, cell, width_cm=None, height_cm=None):
        """
        Inserta una imagen en una celda específica del worksheet.
        Si se proporcionan width_cm y height_cm, redimensiona la imagen.
        """
        if not img_bytes:
            return

        if isinstance(img_bytes, dict):
            img_bytes = img_bytes.get('bytes')
        if not img_bytes:
            return

        try:
            pil_img = Image.open(io.BytesIO(img_bytes))
            img_buf = io.BytesIO()
            pil_img.save(img_buf, format='PNG')
            img_buf.seek(0)

            xl_img = OpenpyxlImage(img_buf)

            # Redimensionar si se especifica
            if width_cm and height_cm:
                xl_img.width = width_cm / 2.54 * 72   # cm → puntos (72 puntos por pulgada)
                xl_img.height = height_cm / 2.54 * 72
            elif width_cm:
                # Mantener relación de aspecto
                orig_w, orig_h = pil_img.size
                ratio = orig_h / orig_w
                xl_img.width = width_cm / 2.54 * 72
                xl_img.height = xl_img.width * ratio

            ws.add_image(xl_img, cell)
        except Exception as e:
            print(f"Error insertando imagen en {cell}: {e}")

    def _insert_image_from_file(self, ws, filepath, cell, width_cm=None, height_cm=None):
        """
        Inserta una imagen desde un archivo en disco en una celda del worksheet.
        Se usa para fotos de moto y piloto.
        """
        if not filepath or not os.path.exists(filepath):
            return

        try:
            pil_img = Image.open(filepath)
            img_buf = io.BytesIO()
            pil_img.save(img_buf, format='PNG')
            img_buf.seek(0)

            xl_img = OpenpyxlImage(img_buf)

            if width_cm and height_cm:
                xl_img.width = width_cm / 2.54 * 72
                xl_img.height = height_cm / 2.54 * 72
            elif width_cm:
                orig_w, orig_h = pil_img.size
                ratio = orig_h / orig_w
                xl_img.width = width_cm / 2.54 * 72
                xl_img.height = xl_img.width * ratio

            ws.add_image(xl_img, cell)
        except Exception as e:
            print(f"Error insertando imagen de archivo {filepath} en {cell}: {e}")

    @staticmethod
    def _fmt(val, dec=2):
        """Formatea un valor numérico."""
        if val is None or val == "":
            return ""
        try:
            return round(float(val), dec)
        except (ValueError, TypeError):
            return str(val)

    @staticmethod
    def _fmt_unit(val, unit, dec=0):
        """Formatea un valor numérico y le añade la unidad (ej: '80 kg', '175 cm')."""
        if val is None or val == "" or val == 0 or val == "0":
            return ""
        try:
            v = round(float(val), dec)
            if dec == 0:
                return f"{int(v)} {unit}"
            return f"{v} {unit}"
        except (ValueError, TypeError):
            return str(val)

    def generate_accel_recovery(self, preview_data):
        """
        Genera el reporte Excel de aceleración y recuperación
        usando el formato ft-nm-000-008.xlsx.
        """
        try:
            template_path = os.path.join(self.templates_dir, "FT-NM-000-008V1.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # Leer mapeo de celdas del módulo
            cells = preview_data.get('excel_cells', {})
            sizes = preview_data.get('excel_img_sizes', {})

            moto = preview_data.get('moto_info', {})
            env_cond = preview_data.get('env_conditions', {})
            lugar = env_cond.get('lugar', {}) if env_cond else {}
            inputs = preview_data.get('inputs', [{}])[0]
            ctx = preview_data.get('contexto_gps', {})

            # ── INFORMACIÓN GENERAL ──
            ws[cells.get("modelo_codigo", "B9")] = f"{moto.get('Nombre Comercial', '')} ({moto.get('Código Modelo', '')})"
            ws[cells.get("fecha", "N7")] = pd.Timestamp.now().strftime("%d/%m/%Y")
            ws[cells.get("placa", "N9")] = moto.get('Placa', '')
            ws[cells.get("chasis", "N10")] = moto.get('Chasis', '')
            ws[cells.get("motor", "N11")] = moto.get('Motor', '')
            ws[cells.get("origen", "N8")] = moto.get('Origen', '')
            ws[cells.get("comentarios", "A24")] = preview_data.get('comments', '')

            # ── FOTO DE MOTO ──
            from ui.management.motos_view import get_moto_foto_path
            moto_foto = get_moto_foto_path(moto)
            if moto_foto:
                self._insert_image_from_file(ws, moto_foto,
                                            cells.get("foto_moto", "Y7"),
                                            width_cm=5.0)

            # ── CONDICIONES ──
            ws[cells.get("piloto_nombre", "B30")] = inputs.get('pilot', '')
            ws[cells.get("piloto_peso", "B31")] = self._fmt_unit(inputs.get('weight', ''), 'kg')
            ws[cells.get("piloto_altura", "B32")] = self._fmt_unit(inputs.get('altura', ''), 'cm')

            # ── FOTO DE PILOTO ──
            from ui.management.pilotos_view import get_piloto_foto_path
            piloto_foto = get_piloto_foto_path(inputs.get('pilot', ''))
            if piloto_foto:
                self._insert_image_from_file(ws, piloto_foto,
                                            cells.get("piloto_foto", "B34"),
                                            width_cm=5.0)

            # ── CONDICIONES - PASAJERO SI EXISTE ──
            pax_name = inputs.get('passenger', '')
            if pax_name:
                ws[cells.get("pasajero_nombre", "J46")] = pax_name
                ws[cells.get("pasajero_peso", "J47")] = self._fmt_unit(inputs.get('pax_weight', ''), 'kg')
                ws[cells.get("pasajero_altura", "J48")] = self._fmt_unit(inputs.get('pax_altura', ''), 'cm')
                pax_foto = get_piloto_foto_path(pax_name)
                if pax_foto:
                    self._insert_image_from_file(ws, pax_foto,
                                                cells.get("pasajero_foto", "J50"),
                                                width_cm=5.0)

            ws[cells.get("lugar_nombre", "W30")] = lugar.get('Nombre', '')
            if ctx:
                ws[cells.get("lugar_altitud", "W31")] = self._fmt(ctx.get('altitud_promedio_msnm', ''))
                ws[cells.get("lugar_coordenadas", "W32")] = f"{ctx.get('latitud_inicial', '')}, {ctx.get('longitud_inicial', '')}"
                ws[cells.get("lugar_link", "W33")] = ctx.get('google_maps_link', '')
            else:
                ws[cells.get("lugar_altitud", "W31")] = lugar.get('Altitud (msnm)', '')

            ws[cells.get("temp_ambiente", "AG30")] = self._fmt(env_cond.get('temp_amb', '')) if env_cond else ''
            ws[cells.get("humedad", "AG31")] = self._fmt(env_cond.get('humidity', '')) if env_cond else ''
            ws[cells.get("temp_suelo", "AG32")] = self._fmt(env_cond.get('temp_ground', '')) if env_cond else ''

            # ── MAPA CONTEXTO ──
            map_size = sizes.get("mapa", (10.5, 15.5))
            self._insert_image(ws, preview_data.get('context_map'),
                              cells.get("mapa_trazado", "T35"),
                              width_cm=map_size[1], height_cm=map_size[0])

            # ── ACELERACIÓN 0-80 ──
            a_data = preview_data.get('accel_data')
            if a_data:
                start_row = cells.get("accel_start_row", 59)
                for i, ev in enumerate(a_data.get('top_3_events', [])):
                    row = start_row + i
                    m = ev['metrics']
                    ws[f"{cells.get('accel_col_num', 'B')}{row}"] = i + 1
                    ws[f"{cells.get('accel_col_evento', 'C')}{row}"] = ev.get('display_name', f"Event {ev['id']}")
                    ws[f"{cells.get('accel_col_vi', 'E')}{row}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{cells.get('accel_col_vf', 'G')}{row}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{cells.get('accel_col_tiempo', 'I')}{row}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{cells.get('accel_col_dist', 'K')}{row}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{cells.get('accel_col_acel', 'M')}{row}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{cells.get('accel_col_rpm', 'O')}{row}"] = self._fmt(m.get('top_rpm', 0), 0)

                # Gráfica resumen
                res_size = sizes.get("grafica_resumen", (11.5, 17.5))
                self._insert_image(ws, a_data.get('img_combined'),
                                  cells.get("accel_img_resumen", "R55"),
                                  width_cm=res_size[1], height_cm=res_size[0])

                # Best event - Tabla de segmentos (filas 101-105)
                seg_start = cells.get("seg_start_row", 101)
                segments = a_data.get('segments', [])
                for i, seg in enumerate(segments):
                    row = seg_start + i
                    # seg = [segment_name, time, distance, acc, rpm]
                    ws[f"{cells.get('seg_col_name', 'B')}{row}"] = seg[0]
                    ws[f"{cells.get('seg_col_time', 'E')}{row}"] = self._fmt(seg[1])
                    ws[f"{cells.get('seg_col_dist', 'H')}{row}"] = self._fmt(seg[2])
                    ws[f"{cells.get('seg_col_acc', 'K')}{row}"] = self._fmt(seg[3])
                    ws[f"{cells.get('seg_col_rpm', 'N')}{row}"] = self._fmt(seg[4], 0)

                # Imágenes del mejor evento
                vel_size = sizes.get("grafica_vel", (7.0, 17.5))
                small_size = sizes.get("grafica_small", (4.0, 17.5))

                self._insert_image(ws, a_data.get('img_detail_gps'),
                                  cells.get("best_accel_img_mapa", "B107"),
                                  width_cm=map_size[1], height_cm=map_size[0])
                self._insert_image(ws, a_data.get('img_detail_v'),
                                  cells.get("best_accel_img_vel", "R97"),
                                  width_cm=vel_size[1], height_cm=vel_size[0])
                self._insert_image(ws, a_data.get('img_detail_a'),
                                  cells.get("best_accel_img_acel", "R109"),
                                  width_cm=small_size[1], height_cm=small_size[0])
                self._insert_image(ws, a_data.get('img_detail_rpm'),
                                  cells.get("best_accel_img_rpm", "R115"),
                                  width_cm=small_size[1], height_cm=small_size[0])

            # ── RECUPERACIÓN ──
            r_data = preview_data.get('recovery_data')
            if r_data:
                # Tabla resumen
                rec_start = cells.get("rec_start_row", 79)
                for i, ev in enumerate(r_data.get('summary_events', [])):
                    row = rec_start + i
                    m = ev['metrics']
                    ws[f"{cells.get('rec_col_num', 'B')}{row}"] = i + 1
                    ws[f"{cells.get('rec_col_evento', 'C')}{row}"] = ev.get('display_name', f"Event {ev['id']}")
                    ws[f"{cells.get('rec_col_vi', 'E')}{row}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{cells.get('rec_col_vf', 'G')}{row}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{cells.get('rec_col_tiempo', 'I')}{row}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{cells.get('rec_col_dist', 'K')}{row}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{cells.get('rec_col_acel', 'M')}{row}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{cells.get('rec_col_rpm', 'O')}{row}"] = self._fmt(m.get('top_rpm', 0), 0)

                res_size = sizes.get("grafica_resumen", (11.5, 17.5))
                self._insert_image(ws, r_data.get('summary_img'),
                                  cells.get("rec_img_resumen", "R75"),
                                  width_cm=res_size[1], height_cm=res_size[0])

                # Bandas individuales
                bands_config = {
                    30: {"row_key": "best_rec30_row", "vel_key": "best_rec30_img_vel",
                         "acel_key": "best_rec30_img_acel", "rpm_key": "best_rec30_img_rpm",
                         "mapa_key": "best_rec30_img_mapa"},
                    40: {"row_key": "best_rec40_row", "vel_key": "best_rec40_img_vel",
                         "acel_key": "best_rec40_img_acel", "rpm_key": "best_rec40_img_rpm",
                         "mapa_key": "best_rec40_img_mapa"},
                    50: {"row_key": "best_rec50_row", "vel_key": "best_rec50_img_vel",
                         "acel_key": "best_rec50_img_acel", "rpm_key": "best_rec50_img_rpm",
                         "mapa_key": "best_rec50_img_mapa"},
                }

                bands = r_data.get('bands', {})
                vel_size = sizes.get("grafica_vel", (7.0, 17.5))
                small_size = sizes.get("grafica_small", (4.0, 17.5))

                for spd, conf in bands_config.items():
                    b_info = bands.get(spd)
                    if not b_info:
                        continue

                    m = b_info['best_event']['metrics']
                    rw = cells.get(conf["row_key"], 126)

                    ws[f"{cells.get('rec_col_num', 'B')}{rw}"] = 1
                    ws[f"{cells.get('rec_col_evento', 'C')}{rw}"] = f"Best {spd}-80"
                    ws[f"{cells.get('rec_col_vi', 'E')}{rw}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{cells.get('rec_col_vf', 'G')}{rw}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{cells.get('rec_col_tiempo', 'I')}{rw}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{cells.get('rec_col_dist', 'K')}{rw}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{cells.get('rec_col_acel', 'M')}{rw}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{cells.get('rec_col_rpm', 'O')}{rw}"] = self._fmt(m.get('top_rpm', 0), 0)

                    self._insert_image(ws, b_info.get('img_gps'),
                                      cells.get(conf["mapa_key"], "B128"),
                                      width_cm=map_size[1], height_cm=map_size[0])
                    self._insert_image(ws, b_info.get('img_v'),
                                      cells.get(conf["vel_key"], "R122"),
                                      width_cm=vel_size[1], height_cm=vel_size[0])
                    self._insert_image(ws, b_info.get('img_a'),
                                      cells.get(conf["acel_key"], "R134"),
                                      width_cm=small_size[1], height_cm=small_size[0])
                    self._insert_image(ws, b_info.get('img_rpm'),
                                      cells.get(conf["rpm_key"], "R140"),
                                      width_cm=small_size[1], height_cm=small_size[0])

            # ── GUARDAR ──
            def clean(s):
                return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()

            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Aceleracion_y_Recuperacion_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            wb.save(filepath)
            return True, filepath

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

    def _fill_common_data(self, ws, preview_data):
        """Llena información general, condiciones y mapa comunes a todos los reportes."""
        cells = preview_data.get('excel_cells', {})
        sizes = preview_data.get('excel_img_sizes', {})
        moto = preview_data.get('moto_info', {})
        env_cond = preview_data.get('env_conditions', {})
        lugar = env_cond.get('lugar', {}) if env_cond else {}
        inputs = preview_data.get('inputs', [{}])
        primary = inputs[0] if inputs else {}
        ctx = preview_data.get('contexto_gps', {})

        # Información general
        ws[cells.get("modelo_codigo", "B9")] = f"{moto.get('Nombre Comercial', '')} ({moto.get('Código Modelo', '')})"
        ws[cells.get("fecha", "N7")] = pd.Timestamp.now().strftime("%d/%m/%Y")
        ws[cells.get("placa", "N9")] = moto.get('Placa', '')
        ws[cells.get("chasis", "N10")] = moto.get('Chasis', '')
        ws[cells.get("motor", "N11")] = moto.get('Motor', '')
        ws[cells.get("origen", "N8")] = moto.get('Origen', '')
        ws[cells.get("comentarios", "A24")] = preview_data.get('comments', '')

        # Foto moto
        from ui.management.motos_view import get_moto_foto_path
        moto_foto = get_moto_foto_path(moto)
        if moto_foto:
            self._insert_image_from_file(ws, moto_foto, cells.get("foto_moto", "Y7"), width_cm=5.0)

        # Condiciones - Piloto
        ws[cells.get("piloto_nombre", "B30")] = primary.get('pilot', '')
        ws[cells.get("piloto_peso", "B31")] = self._fmt_unit(primary.get('weight', ''), 'kg')
        ws[cells.get("piloto_altura", "B32")] = self._fmt_unit(primary.get('altura', ''), 'cm')

        from ui.management.pilotos_view import get_piloto_foto_path
        piloto_foto = get_piloto_foto_path(primary.get('pilot', ''))
        if piloto_foto:
            self._insert_image_from_file(ws, piloto_foto, cells.get("piloto_foto", "B34"), width_cm=5.0)

        # Condiciones - Pasajero si existe
        pax = None
        if len(inputs) > 1 and inputs[1].get('passenger'):
            pax = inputs[1]
        elif inputs and inputs[0].get('passenger'):
            pax = inputs[0]

        if pax:
            ws[cells.get("pasajero_nombre", "J44")] = pax.get('passenger', '')
            ws[cells.get("pasajero_peso", "J45")] = self._fmt_unit(pax.get('pax_weight', ''), 'kg')
            ws[cells.get("pasajero_altura", "J46")] = self._fmt_unit(pax.get('pax_altura', ''), 'cm')
            pax_name = pax.get('passenger', '')
            if pax_name:
                pax_foto = get_piloto_foto_path(pax_name)
                if pax_foto:
                    self._insert_image_from_file(ws, pax_foto, cells.get("pasajero_foto", "J48"), width_cm=5.0)

        # Condiciones - Lugar
        ws[cells.get("lugar_nombre", "W30")] = lugar.get('Nombre', '')
        if ctx:
            ws[cells.get("lugar_altitud", "W31")] = self._fmt(ctx.get('altitud_promedio_msnm', ''))
            ws[cells.get("lugar_coordenadas", "W32")] = f"{ctx.get('latitud_inicial', '')}, {ctx.get('longitud_inicial', '')}"
            ws[cells.get("lugar_link", "W33")] = ctx.get('google_maps_link', '')
        else:
            ws[cells.get("lugar_altitud", "W31")] = lugar.get('Altitud (msnm)', '')

        # Condiciones - Ambiente
        ws[cells.get("temp_ambiente", "AG30")] = self._fmt(env_cond.get('temp_amb', '')) if env_cond else ''
        ws[cells.get("humedad", "AG31")] = self._fmt(env_cond.get('humidity', '')) if env_cond else ''
        ws[cells.get("temp_suelo", "AG32")] = self._fmt(env_cond.get('temp_ground', '')) if env_cond else ''

        # Mapa contexto
        map_size = sizes.get("mapa", (10.5, 15.5))
        self._insert_image(ws, preview_data.get('context_map'),
                          cells.get("mapa_trazado", "T35"),
                          width_cm=map_size[1], height_cm=map_size[0])

        return cells, sizes, map_size

    def generate_braking(self, preview_data):
        """Genera reporte Excel de frenado usando ft-nm-000-005.xlsx."""
        try:
            template_path = os.path.join(self.templates_dir, "FT-NM-000-005V3.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            cells, sizes, map_size = self._fill_common_data(ws, preview_data)

            b_data = preview_data.get('braking_data', {})
            vel_size = sizes.get("grafica_vel", (7.0, 17.5))
            small_size = sizes.get("grafica_small", (4.0, 17.5))
            res_size = sizes.get("grafica_resumen", (11.5, 17.5))

            speed_configs = {
                40: {
                    "start_row": cells.get("brake40_start_row", 72),
                    "col_num": cells.get("brake40_col_num", "B"),
                    "col_evento": cells.get("brake40_col_evento", "C"),
                    "col_vi": cells.get("brake40_col_vi", "E"),
                    "col_vf": cells.get("brake40_col_vf", "G"),
                    "col_tiempo": cells.get("brake40_col_tiempo", "I"),
                    "col_dist": cells.get("brake40_col_dist", "K"),
                    "col_acel": cells.get("brake40_col_acel", "M"),
                    "col_rpm": cells.get("brake40_col_rpm", "O"),
                    "img_resumen": cells.get("brake40_img_resumen", "R68"),
                    "best_row": cells.get("best_brake40_row", 114),
                    "img_vel": cells.get("best_brake40_img_vel", "R110"),
                    "img_acel": cells.get("best_brake40_img_acel", "R122"),
                    "img_rpm": cells.get("best_brake40_img_rpm", "R128"),
                    "img_mapa": cells.get("best_brake40_img_mapa", "B116"),
                },
                60: {
                    "start_row": cells.get("brake60_start_row", 92),
                    "col_num": cells.get("brake60_col_num", "B"),
                    "col_evento": cells.get("brake60_col_evento", "C"),
                    "col_vi": cells.get("brake60_col_vi", "E"),
                    "col_vf": cells.get("brake60_col_vf", "G"),
                    "col_tiempo": cells.get("brake60_col_tiempo", "I"),
                    "col_dist": cells.get("brake60_col_dist", "K"),
                    "col_acel": cells.get("brake60_col_acel", "M"),
                    "col_rpm": cells.get("brake60_col_rpm", "O"),
                    "img_resumen": cells.get("brake60_img_resumen", "R88"),
                    "best_row": cells.get("best_brake60_row", 139),
                    "img_vel": cells.get("best_brake60_img_vel", "R135"),
                    "img_acel": cells.get("best_brake60_img_acel", "R147"),
                    "img_rpm": cells.get("best_brake60_img_rpm", "R153"),
                    "img_mapa": cells.get("best_brake60_img_mapa", "B141"),
                },
            }

            for spd, conf in speed_configs.items():
                spd_data = b_data.get(spd)
                if not spd_data:
                    continue

                # Tabla mejores eventos
                for i, ev in enumerate(spd_data.get('top_3_events', [])):
                    row = conf["start_row"] + i
                    m = ev['metrics']
                    ws[f"{conf['col_num']}{row}"] = i + 1
                    ws[f"{conf['col_evento']}{row}"] = ev.get('display_name', f"Event {ev['id']}")
                    ws[f"{conf['col_vi']}{row}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{conf['col_vf']}{row}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{conf['col_tiempo']}{row}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{conf['col_dist']}{row}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{conf['col_acel']}{row}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{conf['col_rpm']}{row}"] = self._fmt(m.get('top_rpm', 0), 0)

                # Gráfica resumen
                self._insert_image(ws, spd_data.get('img_combined'), conf["img_resumen"],
                                  width_cm=res_size[1], height_cm=res_size[0])

                # Best event
                best = spd_data.get('best_event')
                if best:
                    rw = conf["best_row"]
                    m = best['metrics']
                    ws[f"{conf['col_num']}{rw}"] = 1
                    ws[f"{conf['col_evento']}{rw}"] = f"Best {spd}→0"
                    ws[f"{conf['col_vi']}{rw}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{conf['col_vf']}{rw}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{conf['col_tiempo']}{rw}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{conf['col_dist']}{rw}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{conf['col_acel']}{rw}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{conf['col_rpm']}{rw}"] = self._fmt(m.get('top_rpm', 0), 0)

                    self._insert_image(ws, spd_data.get('img_detail_gps'), conf["img_mapa"],
                                      width_cm=map_size[1], height_cm=map_size[0])
                    self._insert_image(ws, spd_data.get('img_detail_v'), conf["img_vel"],
                                      width_cm=vel_size[1], height_cm=vel_size[0])
                    self._insert_image(ws, spd_data.get('img_detail_a'), conf["img_acel"],
                                      width_cm=small_size[1], height_cm=small_size[0])
                    self._insert_image(ws, spd_data.get('img_detail_rpm'), conf["img_rpm"],
                                      width_cm=small_size[1], height_cm=small_size[0])

            # Guardar
            def clean(s):
                return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto = preview_data.get('moto_info', {})
            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Frenado_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            return True, filepath

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

    def generate_climbing(self, preview_data):
        """Genera reporte Excel de ascenso usando ft-nm-000-012.xlsx."""
        try:
            template_path = os.path.join(self.templates_dir, "FT-NM-000-012V4.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            cells, sizes, map_size = self._fill_common_data(ws, preview_data)

            c_data = preview_data.get('climbing_data', {})
            vel_size = sizes.get("grafica_vel", (7.0, 17.5))
            small_size = sizes.get("grafica_small", (4.0, 17.5))
            res_size = sizes.get("grafica_resumen", (11.5, 17.5))

            file_configs = {
                1: {  # Solo piloto
                    "start_row": cells.get("climb_pilot_start_row", 73),
                    "col_num": cells.get("climb_pilot_col_num", "B"),
                    "col_evento": cells.get("climb_pilot_col_evento", "C"),
                    "col_vi": cells.get("climb_pilot_col_vi", "E"),
                    "col_vf": cells.get("climb_pilot_col_vf", "G"),
                    "col_tiempo": cells.get("climb_pilot_col_tiempo", "I"),
                    "col_dist": cells.get("climb_pilot_col_dist", "K"),
                    "col_acel": cells.get("climb_pilot_col_acel", "M"),
                    "col_rpm": cells.get("climb_pilot_col_rpm", "O"),
                    "img_resumen": cells.get("climb_pilot_img_resumen", "R69"),
                    "best_row": cells.get("best_climb_pilot_row", 115),
                    "img_vel": cells.get("best_climb_pilot_img_vel", "R111"),
                    "img_acel": cells.get("best_climb_pilot_img_acel", "R123"),
                    "img_rpm": cells.get("best_climb_pilot_img_rpm", "R129"),
                    "img_mapa": cells.get("best_climb_pilot_img_mapa", "B117"),
                },
                2: {  # Piloto + pasajero
                    "start_row": cells.get("climb_pax_start_row", 93),
                    "col_num": cells.get("climb_pax_col_num", "B"),
                    "col_evento": cells.get("climb_pax_col_evento", "C"),
                    "col_vi": cells.get("climb_pax_col_vi", "E"),
                    "col_vf": cells.get("climb_pax_col_vf", "G"),
                    "col_tiempo": cells.get("climb_pax_col_tiempo", "I"),
                    "col_dist": cells.get("climb_pax_col_dist", "K"),
                    "col_acel": cells.get("climb_pax_col_acel", "M"),
                    "col_rpm": cells.get("climb_pax_col_rpm", "O"),
                    "img_resumen": cells.get("climb_pax_img_resumen", "R89"),
                    "best_row": cells.get("best_climb_pax_row", 140),
                    "img_vel": cells.get("best_climb_pax_img_vel", "R136"),
                    "img_acel": cells.get("best_climb_pax_img_acel", "R148"),
                    "img_rpm": cells.get("best_climb_pax_img_rpm", "R154"),
                    "img_mapa": cells.get("best_climb_pax_img_mapa", "B142"),
                },
            }

            for file_num, conf in file_configs.items():
                fn_data = c_data.get(file_num)
                if not fn_data:
                    continue

                for i, ev in enumerate(fn_data.get('top_3_events', [])):
                    row = conf["start_row"] + i
                    m = ev['metrics']
                    ws[f"{conf['col_num']}{row}"] = i + 1
                    ws[f"{conf['col_evento']}{row}"] = ev.get('display_name', f"Event {ev['id']}")
                    ws[f"{conf['col_vi']}{row}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{conf['col_vf']}{row}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{conf['col_tiempo']}{row}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{conf['col_dist']}{row}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{conf['col_acel']}{row}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{conf['col_rpm']}{row}"] = self._fmt(m.get('top_rpm', 0), 0)

                self._insert_image(ws, fn_data.get('img_combined'), conf["img_resumen"],
                                  width_cm=res_size[1], height_cm=res_size[0])

                best = fn_data.get('best_event')
                if best:
                    rw = conf["best_row"]
                    m = best['metrics']
                    ws[f"{conf['col_num']}{rw}"] = 1
                    ws[f"{conf['col_evento']}{rw}"] = "Best"
                    ws[f"{conf['col_vi']}{rw}"] = self._fmt(m.get('v_start', 0))
                    ws[f"{conf['col_vf']}{rw}"] = self._fmt(m.get('v_final', 0))
                    ws[f"{conf['col_tiempo']}{rw}"] = self._fmt(m.get('time_s', 0))
                    ws[f"{conf['col_dist']}{rw}"] = self._fmt(m.get('dist_m', 0))
                    ws[f"{conf['col_acel']}{rw}"] = self._fmt(m.get('avg_acc', 0))
                    ws[f"{conf['col_rpm']}{rw}"] = self._fmt(m.get('top_rpm', 0), 0)

                    self._insert_image(ws, fn_data.get('img_detail_gps'), conf["img_mapa"],
                                      width_cm=map_size[1], height_cm=map_size[0])
                    self._insert_image(ws, fn_data.get('img_detail_v'), conf["img_vel"],
                                      width_cm=vel_size[1], height_cm=vel_size[0])
                    self._insert_image(ws, fn_data.get('img_detail_a'), conf["img_acel"],
                                      width_cm=small_size[1], height_cm=small_size[0])
                    self._insert_image(ws, fn_data.get('img_detail_rpm'), conf["img_rpm"],
                                      width_cm=small_size[1], height_cm=small_size[0])

            # Guardar
            def clean(s):
                return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto = preview_data.get('moto_info', {})
            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Ascenso_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            return True, filepath

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

    def generate_topspeed(self, preview_data):
        """Genera reporte Excel de velocidad máxima usando ft-nm-000-007.xlsx."""
        try:
            template_path = os.path.join(self.templates_dir, "FT-NM-000-007V4.xlsx")
            if not os.path.exists(template_path):
                return False, f"Plantilla no encontrada: {template_path}"

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            cells, sizes, map_size = self._fill_common_data(ws, preview_data)

            ts_data = preview_data.get('topspeed_data', {})
            vel_size = sizes.get("grafica_vel", (7.0, 17.5))
            small_size = sizes.get("grafica_small", (4.0, 17.5))
            res_size = sizes.get("grafica_resumen", (11.5, 17.5))

            # Tabla mejores eventos
            start_row = cells.get("topspeed_start_row", 70)
            for i, ev in enumerate(ts_data.get('top_3_events', [])):
                row = start_row + i
                m = ev['metrics']
                ws[f"{cells.get('topspeed_col_num', 'B')}{row}"] = i + 1
                ws[f"{cells.get('topspeed_col_evento', 'C')}{row}"] = ev.get('display_name', f"Event {ev['id']}")
                ws[f"{cells.get('topspeed_col_vi', 'E')}{row}"] = self._fmt(m.get('v_start', 0))
                ws[f"{cells.get('topspeed_col_vf', 'G')}{row}"] = self._fmt(m.get('max_speed', m.get('v_final', 0)))
                ws[f"{cells.get('topspeed_col_tiempo', 'I')}{row}"] = self._fmt(m.get('time_s', 0))
                ws[f"{cells.get('topspeed_col_dist', 'K')}{row}"] = self._fmt(m.get('dist_m', 0))
                ws[f"{cells.get('topspeed_col_acel', 'M')}{row}"] = self._fmt(m.get('avg_acc', 0))
                ws[f"{cells.get('topspeed_col_rpm', 'O')}{row}"] = self._fmt(m.get('top_rpm', 0), 0)

            # Gráfica resumen
            self._insert_image(ws, ts_data.get('img_combined'),
                              cells.get("topspeed_img_resumen", "R66"),
                              width_cm=res_size[1], height_cm=res_size[0])

            # Best event
            best = ts_data.get('best_event')
            if best:
                m = best['metrics']
                ws[cells.get("best_topspeed_vel_max", "B92")] = self._fmt(m.get('max_speed', 0))
                ws[cells.get("best_topspeed_acel", "K92")] = self._fmt(m.get('avg_acc', 0))
                ws[cells.get("best_topspeed_rpm", "N92")] = self._fmt(m.get('top_rpm', 0), 0)

                # Velocímetro del tablero
                dash = ts_data.get('dashboard_speed')
                if dash is not None:
                    ws[cells.get("dashboard_speed", "E92")] = self._fmt(dash, 1)
                    diff = ts_data.get('speed_diff')
                    if diff is not None:
                        ws[cells.get("speed_diff", "H92")] = self._fmt(diff)

                self._insert_image(ws, ts_data.get('img_detail_gps'),
                                  cells.get("best_topspeed_img_mapa", "B94"),
                                  width_cm=map_size[1], height_cm=map_size[0])
                self._insert_image(ws, ts_data.get('img_detail_v'),
                                  cells.get("best_topspeed_img_vel", "R88"),
                                  width_cm=vel_size[1], height_cm=vel_size[0])
                self._insert_image(ws, ts_data.get('img_detail_a'),
                                  cells.get("best_topspeed_img_acel", "R100"),
                                  width_cm=small_size[1], height_cm=small_size[0])
                self._insert_image(ws, ts_data.get('img_detail_rpm'),
                                  cells.get("best_topspeed_img_rpm", "R106"),
                                  width_cm=small_size[1], height_cm=small_size[0])

            # Guardar
            def clean(s):
                return "".join([c for c in str(s) if c.isalnum() or c in (' ', '-', '_')]).strip()
            moto = preview_data.get('moto_info', {})
            moto_str = clean(moto.get('Nombre Comercial', 'Moto'))
            fecha_str = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Velocidad_Maxima_{moto_str}_{fecha_str}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            return True, filepath

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

